def GetExpectancyOfBeingDownstreamTarget(inhibitionThreshold, ratioThreshold, probabilityThreshold, cellLinesFiles):
	# cellLines should specify the list of cell lines for which one wants to perform the algorithm. If a cell line is specified in this list, an accompanying excel file must also exist. These excel files must be named with "_all.xlsm" appended to the cell line name, e.g. "MCF7_all.xlsm". The excel file must contain two worksheets; "pvalue", and "fold", which should contain the p-values and fold change data, respectively.
	# To do: the specific format of these excel files should be described somewhere, or perhaps an example file provided.

	# First let's iterate through every kinase and record the compounds for which there is inhibition.
	numKinases = 163
	numCompounds = 61

	# This dictionary will contain all the inhibition values for every kinase in every compound where value < inhibitionThreshold
	kinasesInhibitedCompounds = {}

	import csv
	with open('requiredData/kinaseInhibitionSpecificity.csv',"rt") as combinedKinaseInhibitions:
		reader = csv.reader(combinedKinaseInhibitions)
		combinedKinaseInhibitionsList = list(reader)
	for kinaseRow in combinedKinaseInhibitionsList[1:numKinases+1]:
		kinaseValues = []
		compounds = []
		for i, value in enumerate(kinaseRow[1:numCompounds+1]):
			if(float(value) < inhibitionThreshold):
				kinaseValues.append(float(value))
				compounds.append(combinedKinaseInhibitionsList[0][i+1].split(".")[-1])
		kinasesInhibitedCompounds[kinaseRow[0]] = (compounds,kinaseValues)

	# Store the list of compounds in an interable list. This will be useful for later when we need to hash our dictionary with the compound name.
	listOfCompounds = [compound.split(".")[1] for compound in combinedKinaseInhibitionsList[0][1:]]

	# Since the column headings in the Kuster DiscoverX data and the cell line worksheets are different and not in the same order, it will be useful to provide dictionaries linking the compound to the column heading for each data set. We can create similar dictionaries for each cell line later. This will give us a common link - via the keys in the dictionaries - for compounds between data sets.
	compoundsKuster = {kusterHeading.split(".")[1]: kusterHeading for kusterHeading in combinedKinaseInhibitionsList[0][1:]}

	# We will need to import the list of kinases for each cell line which is store in listOfKinases.csv. We can then store this in a dictionary where the key and value is the cell line and string of kinases, respectively

	# The key value in listOfKinases needs to match that of the input cell line file name. First let's extract the cell line from the cell line file name
	cellLines = []
	for cellLineFile in cellLinesFiles:
		extensionIndex = cellLineFile.find(".xl")
		cellLines.append(cellLineFile[:extensionIndex])

	with open('requiredData/listOfKinases.csv',"rt") as listOfKinases:
		reader = csv.reader(listOfKinases)
		listOfKinasesList = list(reader)
	listOfKinases = {row[0]: row[2] for row in listOfKinasesList[1:]}

	# Check if listOfKinases is correctly populated for all the specified cell lines
	for i, cellLine in enumerate(cellLines):
		if not cellLine in listOfKinases:
			print("Error: No kinases imported for "+cellLine+". Is this cell line specified in listOfKinases.csv?")
			quit()

	# We will need to parse excel files. Let's use openpyxml
	import openpyxl

	# Now we will iterate through each cell line where we will retrieve the fold change and p-values for each phosphosite in each compound. The EBDT algorithm will then be performed per cell line.
	for i,cellLine in enumerate(cellLines):
		print("Working on "+cellLine+" file ["+str(i+1)+"/"+str(len(cellLines))+"]")
		cellLineWb = openpyxl.load_workbook(cellLinesFiles[i], read_only=False, keep_vba=True)

		# Fetch the fold change and p-values for this cell line.
		fValues, pValues, fdrValues, compoundsCellLine, sitesCellLine = getDataForKusterCompounds(cellLineWb, listOfCompounds, numCompounds)
		
		# Now we have the fold change and p-values we can correlate the fold change data with the compound inhibitor specificities.
		ratios = correlatePhosphoPeptideWithInhibitorSpecificity(cellLineWb, kinasesInhibitedCompounds, fValues, pValues, compoundsKuster, compoundsCellLine)
		
		probKinases = getProbabilityofBeingKinaseSubs(cellLineWb, listOfKinases, ratios, cellLine)
		
		kinaseSubstrates = makeKSlistOfKinaseDownstreamTargets(cellLineWb, ratios, probKinases, pValues, fdrValues, ratioThreshold, probabilityThreshold)
		
		makePathways(cellLineWb, probKinases, sitesCellLine)

		createNodeEdges(cellLineWb, kinaseSubstrates)
		
		cellLineWb.save(cellLinesFiles[i])
		
def getDataForKusterCompounds(cellLineWb, listOfCompounds, numCompounds):
	# Whilst we're here we can create the new worksheets called "pvalue.select" and "fold.select"
	if 'pvalue.select' not in cellLineWb.sheetnames:
		cellLineWb.create_sheet('pvalue.select')
	if 'fold.select' not in cellLineWb.sheetnames:
		cellLineWb.create_sheet('fold.select')

	wsFValues = cellLineWb["fold"]

	# Let's create the dictionary which will link the compound names listed in listOfCompounds to the coumn headings in the cell line file
	compoundsCellLine = {cellLineHeading.value.split(".")[1]: cellLineHeading.value for cellLineHeading in list(wsFValues.rows)[0][2:len(listOfCompounds)+2]}

	fValues = {}
	pValues = {}
	fdrValues = {}
	sitesCellLine = [row[0].value for row in list(wsFValues)[1:]]

	# Fill the fValues dictionary with the fold change values.
	compoundHeadings = list(compoundsCellLine.keys())
	for j,row in enumerate(list(wsFValues)[1:]):
		siteDict = {}
		for i, cell in enumerate(row[2:len(compoundsCellLine)+2]):
			siteDict[compoundHeadings[i]] = float(cell.value)
		fValues[row[0].value] = siteDict

	# And now fill the p-value dictionary
	wsPValues = cellLineWb["pvalue"]
	for row in list(wsPValues)[1:]:
		siteDict = {}
		for i, cell in enumerate(row[2:len(compoundsCellLine)+2]):
			siteDict[compoundHeadings[i]] = float(cell.value)
		pValues[row[0].value] = siteDict

	for row in list(wsPValues)[1:]:
		fdrValues[row[0].value] = float(row[1].value or 0)

	return fValues, pValues, fdrValues, compoundsCellLine, sitesCellLine

def correlatePhosphoPeptideWithInhibitorSpecificity(cellLineWb, kinasesInhibitedCompounds, fValues, pValues, compoundsKuster, compoundsCellLine):
	# This function correlates the fold change data for phosphosites with compound inhibitor specificities.
	if 'corrPPwithKinases' not in cellLineWb.sheetnames:
		cellLineWb.create_sheet('corrPPwithKinases')

	# We'll use the corrcoef function from numpy
	import numpy as np

	# We will store the correlation values in this dictionary
	correlations = {}

	for kinaseName, kinaseTuple in kinasesInhibitedCompounds.items():
		if len(kinaseTuple[0]) > 1:
			kinaseCorrDict = {}
			for phosphoSiteName, phosphoSiteDict in fValues.items():
				foldChangeValues = []
				kinaseValues = []
				correlation = 0
				for i, compound in enumerate(kinaseTuple[0]):
					if compound in phosphoSiteDict:
						foldChangeValues.append(phosphoSiteDict[compound])
						kinaseValues.append(kinaseTuple[1][i])
				if len(foldChangeValues) > 1:
					correlation = np.corrcoef(kinaseValues, foldChangeValues)[1,0]
				kinaseCorrDict[phosphoSiteName] = correlation
			correlations[kinaseName] = kinaseCorrDict

	# Now write the correlation information to the excel file
	for i,kinase in enumerate(list(correlations.keys())):
		cellLineWb['corrPPwithKinases'].cell(row=1,column=i+2).value = kinase
	for i,phosphoSite in enumerate(list(fValues.keys())):
		cellLineWb['corrPPwithKinases'].cell(row=i+2,column=1).value = phosphoSite
	for i,kinase in enumerate(list(correlations.keys())):
		for j,phosphoSite in enumerate(list(fValues.keys())):
			cellLineWb['corrPPwithKinases'].cell(row=j+2,column=i+2).value = correlations.get(kinase,0).get(phosphoSite)

	# Now we will calculate the ratio of the number of compounds that inhibit a phosphosite to the number of compounds that inhibit each kinase in vitro
	if 'ratioSigPPoverSignInhSpeci' not in cellLineWb.sheetnames:
		cellLineWb.create_sheet('ratioSigPPoverSignInhSpeci')

	ratios = {}

	for kinaseName, kinaseTuple in kinasesInhibitedCompounds.items():
		kinaseRatioDict = {}
		if len(kinaseTuple[0]) > 1:
			for phosphoSiteName, phosphoSiteDict in fValues.items():
				numCompoundsInhPhosphoSite = 0
				for i, compound in enumerate(kinaseTuple[0]):
					if compound in phosphoSiteDict:
						if phosphoSiteDict[compound] < -1 and pValues[phosphoSiteName][compound] < 0.025:
							numCompoundsInhPhosphoSite += 1
							kinaseRatioDict[phosphoSiteName] = float(numCompoundsInhPhosphoSite/len(kinaseTuple[0]))
						elif not phosphoSiteName in kinaseRatioDict:
							kinaseRatioDict[phosphoSiteName] = 0
		ratios[kinaseName] = kinaseRatioDict

	# Now write the ratio information to the excel file
	for i,kinase in enumerate(list(ratios.keys())):
		cellLineWb['ratioSigPPoverSignInhSpeci'].cell(row=1,column=i+2).value = kinase
	for i,phosphoSite in enumerate(list(fValues.keys())):
		cellLineWb['ratioSigPPoverSignInhSpeci'].cell(row=i+2,column=1).value = phosphoSite
	for i,kinase in enumerate(list(ratios.keys())):
		for j,phosphoSite in enumerate(list(fValues.keys())):
			cellLineWb['ratioSigPPoverSignInhSpeci'].cell(row=j+2,column=i+2).value = ratios.get(kinase,0).get(phosphoSite,0)

	return ratios

def getProbabilityofBeingKinaseSubs(cellLineWb, listOfKinases, ratios, cellLine):

	# First let's get a list of all the kinases in the ratios dictionary that appear in the list of kinases in litOfKinases.csv
	kinasesInCellLine = []
	for kinase in list(ratios.keys()):
		if kinase.find(".") >-1:
			kinaseToFind = kinase[0:kinase.find(".")]
		else:
			kinaseToFind = kinase
		if kinaseToFind in listOfKinases[cellLine]:
			kinasesInCellLine.append(kinase)

	ratiosByPhosphosite = {}
	for kinase in list(ratios.keys()):
		for phosphosite, ratio in ratios[kinase].items():
			if ratiosByPhosphosite.get(phosphosite):
				ratiosByPhosphosite[phosphosite].append(ratio)
			else:
				initialisingList = [ratio]
				ratiosByPhosphosite[phosphosite] = initialisingList
	maxRatios = {}
	for phosphosite, phosphositeList in ratiosByPhosphosite.items():
		maxRatios[phosphosite] = max(phosphositeList)

	probOfBeingKinaseSubstrate = {}
	for kinase, phosphositeDict in ratios.items():
		if kinase in kinasesInCellLine:
			kinaseDict = {}
			for phosphosite, ratio in phosphositeDict.items():
				if maxRatios[phosphosite] > 0:
					kinaseDict[phosphosite] = ratio / maxRatios[phosphosite]
				else:
					kinaseDict[phosphosite] = 0
			probOfBeingKinaseSubstrate[kinase] = kinaseDict
		else:
			probOfBeingKinaseSubstrate[kinase] = {phosphosite:0 for phosphosite in list(phosphositeDict.keys())}

	if 'ProbOfBeingKinaseSubs' not in cellLineWb.sheetnames:
		cellLineWb.create_sheet('ProbOfBeingKinaseSubs')
	for i,kinase in enumerate(list(probOfBeingKinaseSubstrate.keys())):
		cellLineWb['ProbOfBeingKinaseSubs'].cell(row=1,column=i+2).value = kinase
	for i,phosphoSite in enumerate(list(ratiosByPhosphosite.keys())):
		cellLineWb['ProbOfBeingKinaseSubs'].cell(row=i+2,column=1).value = phosphoSite
	for i,kinase in enumerate(list(probOfBeingKinaseSubstrate.keys())):
		for j,phosphoSite in enumerate(list(probOfBeingKinaseSubstrate[kinase].keys())):
			cellLineWb['ProbOfBeingKinaseSubs'].cell(row=j+2,column=i+2).value = probOfBeingKinaseSubstrate.get(kinase,0).get(phosphoSite)

	return probOfBeingKinaseSubstrate

def makeKSlistOfKinaseDownstreamTargets(cellLineWb, ratios, probKinases, pValues, fdrValues, ratioThreshold, probabilityThreshold):
	residuesToCheck = ["None", "(M", "(R", "(K"]
	kinaseSubstrates = {}
	for kinase, kinaseDict in probKinases.items():
		substrates = []
		for phosphosite, prob in kinaseDict.items():
			if not any(residue in phosphosite for residue in residuesToCheck):
				fdr = fdrValues[phosphosite]
				prob = probKinases[kinase][phosphosite]
				ratio = ratios[kinase][phosphosite]
				if prob > probabilityThreshold and ratio > ratioThreshold and fdr < 0.02:
					substrates.extend(site for site in phosphosite.split(";") if len(site)>0)
		kinaseSubstrates[kinase] = substrates

	if 'PutativeKinaseSubstrates' not in cellLineWb.sheetnames:
		cellLineWb.create_sheet('PutativeKinaseSubstrates')
	cellLineWb['PutativeKinaseSubstrates'].cell(row=1,column=1).value = "kinase"
	cellLineWb['PutativeKinaseSubstrates'].cell(row=1,column=2).value = "n"
	cellLineWb['PutativeKinaseSubstrates'].cell(row=1,column=3).value = "substrates"
	for i,kinase in enumerate(list(kinaseSubstrates.keys())):
		cellLineWb['PutativeKinaseSubstrates'].cell(row=i+2,column=2).value = len(kinaseSubstrates[kinase])
		cellLineWb['PutativeKinaseSubstrates'].cell(row=i+2,column=1).value = kinase
		substratesString = ""
		for substrate in kinaseSubstrates[kinase]:
			substratesString += substrate+";"
		cellLineWb['PutativeKinaseSubstrates'].cell(row=i+2,column=3).value = substratesString

	return kinaseSubstrates

def makePathways(cellLineWb, probOfBeingKinaseSubstrate, sitesCellLine):
	#For each phosphosite, find all the kinases which have a probability of 1 (as calculated in getProbabilityofBeingKinaseSubs())
	phosphositePathways = {}
	for kinase, kinaseDict in probOfBeingKinaseSubstrate.items():
		for phosphosite, value in kinaseDict.items():
			if value == 1:
				phosphositePathways.setdefault(phosphosite,[]).append(kinase)

	# Output these pathways to worksheet "site_pathw"
	# For sake of avoiding confusion with the output, the sheet site_pathw will not be created
	#if 'site_pathw' not in cellLineWb.sheetnames:
	#	cellLineWb.create_sheet('site_pathw')
	#cellLineWb['site_pathw'].cell(row=1,column=1).value = "site"
	#cellLineWb['site_pathw'].cell(row=1,column=2).value = "pathway"
	#for i,site in enumerate(sitesCellLine):
	#	cellLineWb['site_pathw'].cell(row=i+2,column=1).value = site
	#	cellLineWb['site_pathw'].cell(row=i+2,column=2).value = myString = "-".join(phosphositePathways.get(site,""))

	uniquePathways = []
	for pathway in [pathway for pathway in [phosphositePathways[phosphosite] for phosphosite in sitesCellLine if phosphositePathways.get(phosphosite,0) is not 0] if len(pathway)>1]:
		found = False
		pathwayString = "-".join(kinase for kinase in pathway)
		for uniquePathway in uniquePathways:
			uniquePathwayString = "-".join(kinase for kinase in uniquePathway)
			if pathwayString in uniquePathwayString: found = True
		if not found: uniquePathways.append(pathway)

	uniquePathways = uniquePathways[::-1]

	pathwaySites = {}
	for pathway in uniquePathways:
		pathway = list(pathway)
		pathwayString = "-".join(kinase for kinase in pathway)
		sites = []
		for site, sitePathway in phosphositePathways.items():
			
			for i in range(len(sitePathway) - len(pathway) + 1):
				if pathway == sitePathway[i:i+len(pathway)]: sites.append(site)
		if len(sites)>1:
			pathwaySites[pathwayString] = sites

	#Output the sites for each pathway to worksheet "pathways"
	# For sake of avoiding confusion with the output, the sheet pathways will not be created
	#if 'pathways' not in cellLineWb.sheetnames:
	#	cellLineWb.create_sheet('pathways')
	#cellLineWb['pathways'].cell(row=1,column=1).value = "pathway"
	#cellLineWb['pathways'].cell(row=1,column=2).value = "m"
	#cellLineWb['pathways'].cell(row=1,column=3).value = "sites"
	#worksheetRow = 2
	#for pathway, sites in pathwaySites.items():
	#	siteString = ";".join(site for site in sites)
	#	cellLineWb['pathways'].cell(row=worksheetRow,column=3).value = siteString
	#	cellLineWb['pathways'].cell(row=worksheetRow,column=2).value = len(sites)
	#	cellLineWb['pathways'].cell(row=worksheetRow,column=1).value = pathway
	#	worksheetRow += 1

def createNodeEdges(cellLineWb, kinaseSubstrates):
	nodesSubtrates = {}
	for kinase in list(kinaseSubstrates.keys()):
		for kinaseToCheck in list(kinaseSubstrates.keys()):
			if kinase < kinaseToCheck:
				substrates = []
				for substrate in kinaseSubstrates[kinase]:
					if substrate in kinaseSubstrates[kinaseToCheck]:
						substrates.append(substrate)
				if len(substrates)>0:
					key = kinase+"."+kinaseToCheck
					nodesSubtrates[key] = substrates

	if 'tableEdgeSubs' not in cellLineWb.sheetnames:
		cellLineWb.create_sheet('tableEdgeSubs')
	for i, kinase in enumerate(list(kinaseSubstrates.keys())):
		cellLineWb['tableEdgeSubs'].cell(row=i+2, column=1).value = kinase
		cellLineWb['tableEdgeSubs'].cell(row=1, column=i+2).value = kinase
		for j, kinaseToCheck in enumerate(list(kinaseSubstrates.keys())):
			if j > i:
				key = kinase+"."+kinaseToCheck
				cellLineWb['tableEdgeSubs'].cell(row=i+2, column=j+2).value = ";".join([str(substrateList) for substrateList in nodesSubtrates.get(key,[])])
				cellLineWb['tableEdgeSubs'].cell(row=j+2, column=i+2).value = ";".join([str(substrateList) for substrateList in nodesSubtrates.get(key,[])])

	if 'nodes.edges' not in cellLineWb.sheetnames:
		cellLineWb.create_sheet('nodes.edges')
		cellLineWb['nodes.edges'].cell(row=1, column=1).value = "edge"
		cellLineWb['nodes.edges'].cell(row=1, column=2).value = "weight"
		cellLineWb['nodes.edges'].cell(row=1, column=3).value = "subs"
	worksheetRow = 2
	for kinases, substrates in nodesSubtrates.items():
		cellLineWb['nodes.edges'].cell(row=worksheetRow, column=1).value = kinases
		cellLineWb['nodes.edges'].cell(row=worksheetRow, column=2).value = len(substrates)
		cellLineWb['nodes.edges'].cell(row=worksheetRow, column=3).value = ";".join(substrates)
		worksheetRow += 1